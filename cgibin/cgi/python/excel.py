import sys, os, json

text = sys.argv[1]

js = json.loads(text)

from inspect import getmembers
from pprint import pprint
pprint(getmembers(js))

'''
file_xls = js['filename']
sheets = js['sheet']
set_params = js['sets']

from openpyxl import load_workbook
wb = load_workbook(os.path.dirname(os.path.abspath(__file__))+'/'+file_xls)

wname = wb.sheetnames[sheets['index']]
sheet_ranges = wb[wname]



for k in set_params:
    print(set_params[k])
    
print(sheet_ranges['C2'].value)
'''